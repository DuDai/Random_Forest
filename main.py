"""
需安装的库文件：
pip install xlrd==1.2.0
pip install pywin32

换源网址：
- 豆瓣：http://pypi.douban.com/simple/
- 中科大：https://pypi.mirrors.ustc.edu.cn/simple/
- 清华：https://pypi.tuna.tsinghua.edu.cn/simple

换源安装，例如：pip install pywin32 -i https://pypi.tuna.tsinghua.edu.cn/simple

运行指南：
    先运行main.py生成初步处理后的表格
    再运行file_write.py精处理文件
"""
import xlrd
import os
import openpyxl
import attribute_clean
import file_write
from openpyxl import load_workbook
import pandas as pd

# 读某单元格内的数据
# 参数解读：get_row(excel的路径，选择的子表格名字，所读行的位置)
def get_cell(file_path, sheet_name, row_location, col_location):
    # 打开excel表格
    filename = xlrd.open_workbook(file_path)
    # 打印excel所有子表格的名字
    form_name = filename.sheet_names()
    # 获取选择子表格的位置
    subtable_location = form_name.index(sheet_name)
    # 选择子表格为操作对象
    sheet = filename.sheet_by_index(subtable_location)
    cell = sheet.cell_value(row_location, col_location)
    return cell

# 获取数据所在路径，并读取每一行数据
cwd = os.getcwd()
dataset_path = os.path.join(cwd, 'data_cut/')
test_path = dataset_path + 'normal_report_12.xlsx'
wb = openpyxl.load_workbook(test_path)
table = wb.worksheets[0]
nrows = table.max_row # nrows==4187
reqNumbers = len(attribute_clean.request_attribute) # 常规reqNumbers==23
wafNumbers = len(attribute_clean.wafinfo_attribute) # wafNumbers==4
# request = get_cell(test_path, 'report', 2,1)
# location = get_cell(test_path, 'report', 2,2)
# wafinfo = get_cell(test_path, 'report', 2,3)

# 制作data_clean.csv表头
data_cleaned_path = dataset_path + 'normal_data_cleaned_12.xlsx'
file_write.make_list(data_cleaned_path)

wb = load_workbook(data_cleaned_path)  # 打开已存在的文件
sheet = wb.active
# 截取属性值
for rows in range(2, nrows):
    # 截取request、location、wafinfo中的属性值，删除标点符号
    request = get_cell(test_path, 'report', rows, 1)
    location = get_cell(test_path, 'report', rows, 2)
    wafinfo = get_cell(test_path, 'report', rows, 3)
    # 截取request的内容并存入xlsx文件
    for num in range(reqNumbers-1): # 0~21
        idx1 = request.find(attribute_clean.request_attribute[num])
        idx2 = request.find(attribute_clean.request_attribute[num+1])
        idx_last = request.find(attribute_clean.request_attribute[reqNumbers-1])
        if idx1 != -1: # 找到了相关属性
            idx1_length = len(attribute_clean.request_attribute[num])
            idx2_length = len(attribute_clean.request_attribute[num+1])
            str1 = request[idx1 + idx1_length: idx2]
            # print(str1)
            str2 = attribute_clean.string_clean('"', '\[', '\]', '\{', '\}', ':', ',', str1)
        else: # 没有找到相关属性值
            str2 = 'null'
        if idx_last != -1: # 找到了server属性
            str3 = request[idx_last+6 : ] # 单独录入最后一个属性的值
            str4 = attribute_clean.string_clean('"', '\[', '\]', '\{', '\}', ':', ',', str3)
        else:
            str4 = 'null'
        sheet.cell(rows+1, num+1, str2)
        sheet.cell(rows+1, reqNumbers, str4)

    # 截取location的内容并存入xlsx文件
    sheet.cell(rows+1, reqNumbers+1, location)

    # 截取wafinfo的内容并存入xls文件
    for num in range(wafNumbers-1): # 0~3
        idx3 = wafinfo.find(attribute_clean.wafinfo_attribute[num])
        idx4 = wafinfo.find(attribute_clean.wafinfo_attribute[num+1])
        idx_last = wafinfo.find(attribute_clean.wafinfo_attribute[wafNumbers-1])
        if idx3 != -1: # 找到了相关属性
            idx3_length = len(attribute_clean.wafinfo_attribute[num])
            idx4_length = len(attribute_clean.wafinfo_attribute[num+1])
            str1 = wafinfo[idx3 + idx3_length: idx4]
            str2 = attribute_clean.string_clean('"', '\[', '\]', '\{', '\}', ':', ',', str1)
        else:
            str2 = 'null'
        if idx_last != -1: # 找到了data属性
            str3 = wafinfo[idx_last+4 : ]
            str4 = attribute_clean.string_clean('"', '\[', '\]', '\{', '\}', ':', ',', str3)
        else:
            str4 = 'null'
        sheet.cell(rows+1, reqNumbers+1+num+1, str2)
        sheet.cell(rows+1, reqNumbers+1+wafNumbers, str4)
    wb.save(data_cleaned_path)
    print(rows)










