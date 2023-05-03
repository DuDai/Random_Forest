'''
对data_final.xlsx文件中的特征进行再次提取，删除无规律的字符串特征值
将ip手动拆分成四列，将分隔符设置为'.'即可
'''
import os
import pandas as pd
import openpyxl

cwd = os.getcwd()
data_path = os.path.join(cwd, './data_final.xlsx')  # 二次清洗后的表格
new_data_path = os.path.join(cwd, './data_final_2.xlsx')  # 存储删除不必要后的数据

# 实现删除不必要列的操作
def delete_col(data_path, new_data_path):
    data = pd.read_excel(data_path)
    # 删除带有中文的列和长度不规律的字符串特征列
    data = data.drop(columns=['url', 'path', 'location', 'User-Agent', 'Accept-Language', 'resume',
                              'location', 'rayid', 'message'])
    data.to_excel(new_data_path, sheet_name='report', index=False)
    return 0

# 对空白的区域填入0
def give_zero(new_data_path):
    data = openpyxl.load_workbook(new_data_path)  # 打开目标Excel文件
    sheet = data.active  # 打开该Excel里对应的sheet
    for k in range(1, sheet.max_column + 1):  # 对第1至X列单元格遍历
        for i in range(1, sheet.max_row + 1):  # 对第1至X行单元格遍历
            if sheet.cell(row=i, column=k).value is None:  # 如果该单元格为空
                sheet.cell(i, k, '0')  # 填入0值
    data.save(new_data_path)
    return 0

# 对method进行标号
def give_method_tag(new_data_path): # GET==1, HEAD==2, POST==3, PATCH==4, DELETE==5, OPTIONS==6, PUT==7
    data = openpyxl.load_workbook(new_data_path)  # 打开目标Excel文件
    sheet = data.active  # 打开该Excel里对应的sheet
    for i in range(1, sheet.max_row + 1):  # 对第1至X行单元格遍历
        if sheet.cell(row=i, column=5).value == 'GET': # 判断格内的method类型
            sheet.cell(i, 5, '1')  # 填入1
        elif sheet.cell(row=i, column=5).value == 'HEAD':
            sheet.cell(i, 5, '2')
        elif sheet.cell(row=i, column=5).value == 'POST':
            sheet.cell(i, 5, '3')
        elif sheet.cell(row=i, column=5).value == 'PATCH':
            sheet.cell(i, 5, '4')
        elif sheet.cell(row=i, column=5).value == 'DELETE':
            sheet.cell(i, 5, '5')
        elif sheet.cell(row=i, column=5).value == 'OPTIONS':
            sheet.cell(i, 5, '6')
        elif sheet.cell(row=i, column=5).value == 'PUT':
            sheet.cell(i, 5, '7')
        print(i)
    data.save(new_data_path)
    return 0

# 给host添加tag
def give_host_tag(new_data_path):
    '''nrii.org.cn==1, www.escience.org.cn==2, share.escience.net.cn==3, api.escience.org.cn==4,
    escience.org.cn==5, www.nrii.org.cn==6, nstr.escience.net.cn==7, international.escience.org.cn==8
    escience.net.cn==9, Name.ws.nrii.org.cn.baidustaticcdn.techpath/hostws.nrii.org.cn==10
    spath/robots.txt/.ssh/known_hostshostnrii.org.cn==11, analytics.escience.net.cn==12
    oauth.escience.org.cn==13, sys.escience.net.cn==14, ws.nrii.org.cn==15
    sg.nrii.org.cn==16, .keypath/localhost.keyhostnrii.org.cn==17,
    sydw.escience.org.cn==18, m.escience.org.cn==19, 124.207.169.208686==20
    cstr.escience.org.cn==21, .keypath/host.keyhostnrii.org.cn==22
    .keypath/host.keyhostescience.org.cn==23, .keypath/localhost.keyhostescience.org.cn==24
    secgate-test.escience.net.cn==25, .keypath/ssl/localhost.keyhostapi.escience.org.cn==26
    Name%7D.cfrp2r5hfhq82ukk0ft0hpacdjocwuwaj.oast.site/a%7Dpath/solr/admin/collectionshostescience.org.cn==27,
    Name%7D.cfrp2r5hfhq82ukk0ft0kcsbh776xncn5.oast.site/a%7Dpath/solr/admin/collectionshostinternational.escience.org.cn==28,
    Name%7D.cfrp2r5hfhq82ukk0ft053764on6pfi5s.oast.site/a%7Dpath/solr/admin/collectionshostnrii.org.cn==29,
    Name%7D.cfrp2r5hfhq82ukk0ft0kz1ijjacs74h6.oast.site/a%7Dpath/solr/admin/collectionshostnstr.escience.net.cn==30,
    Name%7D.cfrp2r5hfhq82ukk0ft0opmh6k98ymk3f.oast.site/a%7Dpath/solr/admin/collectionshostshare.escience.net.cn==31,
    Name%7D.cfrp2r5hfhq82ukk0ft03wzxfcrc31ooq.oast.site/a%7Dpath/solr/admin/collectionshostwww.escience.org.cn==32
    cstr.escience.net.cn==33,
    Name%7D.cfajfsj6q218ic800010zj8yd36p5i7zk.oast.fun%7Dpath/dr/authentication/oauth2/oauth2loginhostnrii.org.cn==34
    .keypath/localhost.keyhostapi.escience.org.cn==35, registry.escience.org.cn==36
    spath/robots.txt/hostshostnrii.org.cn==37, 124.207.169.108087==38
    '''
    data = openpyxl.load_workbook(new_data_path)  # 打开目标Excel文件
    sheet = data.active  # 打开该Excel里对应的sheet
    for i in range(1, sheet.max_row + 1):  # 对第1至X行单元格遍历
        if 'nrii.org.cn' in sheet.cell(row=i, column=6).value: # 判断格内的host类型
            sheet.cell(i, 6, '1')  # 填入1
        elif sheet.cell(row=i, column=6).value == 'www.escience.org.cn':
            sheet.cell(i, 6, '2')
        elif sheet.cell(row=i, column=6).value == 'share.escience.net.cn':
            sheet.cell(i, 6, '3')
        elif sheet.cell(row=i, column=6).value == 'api.escience.org.cn':
            sheet.cell(i, 6, '4')
        elif sheet.cell(row=i, column=6).value == 'escience.org.cn':
            sheet.cell(i, 6, '5')
        elif sheet.cell(row=i, column=6).value == 'www.nrii.org.cn':
            sheet.cell(i, 6, '6')
        elif sheet.cell(row=i, column=6).value == 'nstr.escience.net.cn':
            sheet.cell(i, 6, '7')
        elif sheet.cell(row=i, column=6).value == 'international.escience.org.cn':
            sheet.cell(i, 6, '8')
        elif sheet.cell(row=i, column=6).value == 'escience.net.cn':
            sheet.cell(i, 6, '9')
        elif sheet.cell(row=i, column=6).value == 'Name.ws.nrii.org.cn.baidustaticcdn.techpath/hostws.nrii.org.cn':
            sheet.cell(i, 6, '10')
        elif sheet.cell(row=i, column=6).value == 'spath/robots.txt/.ssh/known_hostshostnrii.org.cn':
            sheet.cell(i, 6, '11')
        elif sheet.cell(row=i, column=6).value == 'analytics.escience.net.cn':
            sheet.cell(i, 6, '12')
        elif sheet.cell(row=i, column=6).value == 'oauth.escience.org.cn':
            sheet.cell(i, 6, '13')
        elif sheet.cell(row=i, column=6).value == 'sys.escience.net.cn':
            sheet.cell(i, 6, '14')
        elif sheet.cell(row=i, column=6).value == 'ws.nrii.org.cn':
            sheet.cell(i, 6, '15')
        elif sheet.cell(row=i, column=6).value == 'sg.nrii.org.cn':
            sheet.cell(i, 6, '16')
        elif sheet.cell(row=i, column=6).value == '.keypath/localhost.keyhostnrii.org.cn':
            sheet.cell(i, 6, '17')
        elif sheet.cell(row=i, column=6).value == 'sydw.escience.org.cn':
            sheet.cell(i, 6, '18')
        elif sheet.cell(row=i, column=6).value == 'm.escience.org.cn':
            sheet.cell(i, 6, '19')
        elif sheet.cell(row=i, column=6).value == '124.207.169.208686':
            sheet.cell(i, 6, '20')
        elif sheet.cell(row=i, column=6).value == 'cstr.escience.org.cn':
            sheet.cell(i, 6, '21')
        elif sheet.cell(row=i, column=6).value == '.keypath/host.keyhostnrii.org.cn':
            sheet.cell(i, 6, '22')
        elif sheet.cell(row=i, column=6).value == '.keypath/host.keyhostescience.org.cn':
            sheet.cell(i, 6, '23')
        elif sheet.cell(row=i, column=6).value == '.keypath/localhost.keyhostescience.org.cn':
            sheet.cell(i, 6, '24')
        elif sheet.cell(row=i, column=6).value == 'secgate-test.escience.net.cn':
            sheet.cell(i, 6, '25')
        elif sheet.cell(row=i, column=6).value == '.keypath/ssl/localhost.keyhostapi.escience.org.cn':
            sheet.cell(i, 6, '26')
        elif sheet.cell(row=i, column=6).value == 'Name%7D.cfrp2r5hfhq82ukk0ft0hpacdjocwuwaj.oast.site/a%7Dpath/solr/admin/collectionshostescience.org.cn':
            sheet.cell(i, 6, '27')
        elif sheet.cell(row=i, column=6).value == 'Name%7D.cfrp2r5hfhq82ukk0ft0kcsbh776xncn5.oast.site/a%7Dpath/solr/admin/collectionshostinternational.escience.org.cn':
            sheet.cell(i, 6, '28')
        elif sheet.cell(row=i, column=6).value == 'Name%7D.cfrp2r5hfhq82ukk0ft053764on6pfi5s.oast.site/a%7Dpath/solr/admin/collectionshostnrii.org.cn':
            sheet.cell(i, 6, '29')
        elif sheet.cell(row=i, column=6).value == 'Name%7D.cfrp2r5hfhq82ukk0ft0kz1ijjacs74h6.oast.site/a%7Dpath/solr/admin/collectionshostnstr.escience.net.cn':
            sheet.cell(i, 6, '30')
        elif sheet.cell(row=i, column=6).value == 'Name%7D.cfrp2r5hfhq82ukk0ft0opmh6k98ymk3f.oast.site/a%7Dpath/solr/admin/collectionshostshare.escience.net.cn':
            sheet.cell(i, 6, '31')
        elif sheet.cell(row=i, column=6).value == 'Name%7D.cfrp2r5hfhq82ukk0ft03wzxfcrc31ooq.oast.site/a%7Dpath/solr/admin/collectionshostwww.escience.org.cn':
            sheet.cell(i, 6, '32')
        elif sheet.cell(row=i, column=6).value == 'cstr.escience.net.cn':
            sheet.cell(i, 6, '33')
        elif sheet.cell(row=i, column=6).value == 'Name%7D.cfajfsj6q218ic800010zj8yd36p5i7zk.oast.fun%7Dpath/dr/authentication/oauth2/oauth2loginhostnrii.org.cn':
            sheet.cell(i, 6, '34')
        elif sheet.cell(row=i, column=6).value == '.keypath/localhost.keyhostapi.escience.org.cn':
            sheet.cell(i, 6, '35')
        elif sheet.cell(row=i, column=6).value == 'registry.escience.org.cn':
            sheet.cell(i, 6, '36')
        elif sheet.cell(row=i, column=6).value == 'spath/robots.txt/hostshostnrii.org.cn':
            sheet.cell(i, 6, '37')
        elif sheet.cell(row=i, column=6).value == '124.207.169.108087':
            sheet.cell(i, 6, '38')
        print(i)
    data.save(new_data_path)
    return 0

def give_proto_tag(new_data_path):
    '''
    HTTP/1.1==1, HTTP/2.0==2, HTTP/1.0==3
    '''
    data = openpyxl.load_workbook(new_data_path)  # 打开目标Excel文件
    sheet = data.active  # 打开该Excel里对应的sheet
    for i in range(1, sheet.max_row + 1):  # 对第1至X行单元格遍历
        if 'HTTP/1.1' in sheet.cell(row=i, column=7).value:  # 判断格内的ptoto类型
            sheet.cell(i, 7, '1')  # 填入1
        elif 'HTTP/2.0' in sheet.cell(row=i, column=7).value:
            sheet.cell(i, 7, '2')
        elif 'HTTP/1.0' in sheet.cell(row=i, column=7).value:
            sheet.cell(i, 7, '3')
        print(i)
    data.save(new_data_path)
    return 0

def give_encoding_tag(new_data_path):
    '''
    gzip==1, identity==2, deflate==3, null==0
    '''
    data = openpyxl.load_workbook(new_data_path)  # 打开目标Excel文件
    sheet = data.active  # 打开该Excel里对应的sheet
    for i in range(1, sheet.max_row + 1):  # 对第1至X行单元格遍历
        if 'gzip' in sheet.cell(row=i, column=9).value:  # 判断格内的encoding类型
            sheet.cell(i, 9, '1')  # 填入1
        elif 'identity' in sheet.cell(row=i, column=9).value:  # 判断格内的encoding类型
            sheet.cell(i, 9, '2')  # 填入2
        elif 'deflate' in sheet.cell(row=i, column=9).value:  # 判断格内的encoding类型
            sheet.cell(i, 9, '3')
        elif 'null' in sheet.cell(row=i, column=9).value:  # 判断格内的encoding类型
            sheet.cell(i, 9, '0')
        print(i)
    data.save(new_data_path)
    return 0

def give_version_tag(new_data_path):
    '''
    TLS12==1, TLS13==2
    '''
    data = openpyxl.load_workbook(new_data_path)  # 打开目标Excel文件
    sheet = data.active  # 打开该Excel里对应的sheet
    for i in range(1, sheet.max_row + 1):  # 对第1至X行单元格遍历
        if 'TLS12' in sheet.cell(row=i, column=10).value:  # 判断格内的version类型
            sheet.cell(i, 10, '1')  # 填入1
        elif 'TLS13' in sheet.cell(row=i, column=10).value:  # 判断格内的version类型
            sheet.cell(i, 10, '2')  # 填入2
        print(i)
    data.save(new_data_path)
    return 0

def give_ciphersuite_tag(new_data_path):
    '''
    GCM_SHA256==1, unknown 1301==2, CBC_SHA==3, unknown cca8==4, unknown 1303==5, unknown c027==6
    '''
    data = openpyxl.load_workbook(new_data_path)  # 打开目标Excel文件
    sheet = data.active  # 打开该Excel里对应的sheet
    for i in range(1, sheet.max_row + 1):  # 对第1至X行单元格遍历
        if 'GCM_SHA256' in sheet.cell(row=i, column=11).value:
            sheet.cell(i, 11, '1')  # 填入1
        elif 'unknown 1301' in sheet.cell(row=i, column=11).value:
            sheet.cell(i, 11, '2')  # 填入2
        elif 'CBC_SHA' in sheet.cell(row=i, column=11).value:
            sheet.cell(i, 11, '3')
        elif 'unknown cca8' in sheet.cell(row=i, column=11).value:
            sheet.cell(i, 11, '4')
        elif 'unknown 1303' in sheet.cell(row=i, column=11).value:
            sheet.cell(i, 11, '5')
        elif 'unknown c027' in sheet.cell(row=i, column=11).value:
            sheet.cell(i, 11, '6')
        print(i)
    data.save(new_data_path)
    return 0

def give_server_tag(new_data_path):
    '''
    nrii.org.cn==1, escience.org.cn==2, www.escience.org.cn==3, international.escience.org.cn==4, ws.nrii.org.cn==5
    escience.net.cn==6, else==7
    '''
    data = openpyxl.load_workbook(new_data_path)  # 打开目标Excel文件
    sheet = data.active  # 打开该Excel里对应的sheet
    for i in range(2, sheet.max_row + 1):  # 对第1至X行单元格遍历
        if 'ws.nrii.org.cn' in sheet.cell(row=i, column=12).value:  # 判断格内的server类型
            sheet.cell(i, 12, '5')  # 填入5
        elif 'international.escience.org.cn' in sheet.cell(row=i, column=12).value:
            sheet.cell(i, 12, '4')  # 填入4
        elif 'www.escience.org.cn' in sheet.cell(row=i, column=12).value:
            sheet.cell(i, 12, '3')  # 填入3
        elif 'escience.org.cn' in sheet.cell(row=i, column=12).value:
            sheet.cell(i, 12, '2')  # 填入2
        elif 'nrii.org.cn' in sheet.cell(row=i, column=12).value:
            sheet.cell(i, 12, '1')  # 填入1
        elif 'escience.net.cn' in sheet.cell(row=i, column=12).value:
            sheet.cell(i, 12, '6')  # 填入6
        else:
            sheet.cell(i, 12, '7')
        print(i)
    data.save(new_data_path)
    return 0

def give_data_tag(new_data_path):
    '''
    GET==1, Matched Data==2, application==3, PUT==4, PATCH==5
    DELETE==6, invalid JSON==7, invalid character==8, syntax error==9, REQUEST_HEADERS==10, else==11
    '''
    data = openpyxl.load_workbook(new_data_path)  # 打开目标Excel文件
    sheet = data.active  # 打开该Excel里对应的sheet
    for i in range(2, sheet.max_row + 1):  # 对第1至X行单元格遍历
        if 'GET' in sheet.cell(row=i, column=14).value:  # 判断格内的data类型
            sheet.cell(i, 14, '1')
        elif 'Matched Data' in sheet.cell(row=i, column=14).value:
            sheet.cell(i, 14, '2')
        elif 'application' in sheet.cell(row=i, column=14).value:
            sheet.cell(i, 14, '3')
        elif 'PUT' in sheet.cell(row=i, column=14).value:
            sheet.cell(i, 14, '4')
        elif 'PATCH' in sheet.cell(row=i, column=14).value:
            sheet.cell(i, 14, '5')
        elif 'DELETE' in sheet.cell(row=i, column=14).value:
            sheet.cell(i, 14, '6')
        elif 'invalid JSON' in sheet.cell(row=i, column=14).value:
            sheet.cell(i, 14, '7')
        elif 'invalid character' in sheet.cell(row=i, column=14).value:
            sheet.cell(i, 14, '8')
        elif 'syntax error' in sheet.cell(row=i, column=14).value:
            sheet.cell(i, 14, '9')
        elif 'REQUEST_HEADERS' in sheet.cell(row=i, column=14).value:
            sheet.cell(i, 14, '10')
        else:
            sheet.cell(i, 14, '11')
        print(i)
    data.save(new_data_path)
    return 0

# delete_col(data_path, new_data_path)
# give_zero(new_data_path)
give_method_tag(new_data_path)
give_host_tag(new_data_path)
give_proto_tag(new_data_path)
give_encoding_tag(new_data_path)
give_version_tag(new_data_path)
give_ciphersuite_tag(new_data_path)
give_server_tag(new_data_path)
give_data_tag(new_data_path)