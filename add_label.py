'''
为attack_data_all_cleaned.xlsx和normal_data_all_cleaned.xlsx数据分别加标签
攻击样本标签为1，正常样本标签为0
'''
import openpyxl
import os

cwd = os.getcwd()
# attack_data_path = os.path.join(cwd, 'attack_data/attack_data_all_cleaned.xlsx') # 被攻击样本的地址
normal_data_path = os.path.join(cwd, 'normal_data/normal_data_all_cleaned_12.xlsx')  # 正常样本的地址
data_normal = openpyxl.load_workbook(normal_data_path)  # 打开正常的xlsx文件
# data_attack = openpyxl.load_workbook(attack_data_path)       # 打开被攻击样本的xlsx文件
sheet_normal = data_normal.active  # 打开该Excel里对应的sheet
# sheet_attack = data_attack.active

# 添加攻击样本的标签
# attack_data_final = os.path.join(cwd, 'attack_data/attack_data_final.xlsx')
# attack_max_col = sheet_attack.max_column
# # max_col = sheet_attack.max_column
# for i in range(1, sheet_attack.max_row+1):
#     if i == 1:
#         sheet_attack.cell(i, attack_max_col+1, 'attack')
#     else:
#         sheet_attack.cell(i, attack_max_col+1, '1')
#     print(i)
# print(attack_max_col+1)
# data_attack.save(attack_data_final)

# 添加正常样本的标签
normal_data_final = os.path.join(cwd, 'normal_data/normal_data_final_12.xlsx')
normal_max_col = sheet_normal.max_column
# max_col = sheet_attack.max_column
for i in range(1, sheet_normal.max_row+1):
    if i == 1:
        sheet_normal.cell(i, normal_max_col+1, 'attack')
    else:
        sheet_normal.cell(i, normal_max_col+1, '0')
    print(i)
print(normal_max_col+1)
data_normal.save(normal_data_final)
