'''
实现文件表头的插入
实现删除不必要的行或列
实现表中空白内容的填充
新建文件采用xlsxwriter方法，打开已有文件采用load_workbook方法
'''
import pandas as pd
import os
import attribute_clean
import openpyxl
import xlsxwriter

# 制作属性表头
def make_list(data_cleaned_path):
        wb = xlsxwriter.Workbook(data_cleaned_path)  # 新建一个实例
        sheet = wb.add_worksheet('report')  # 工作簿名称
        col1 = attribute_clean.request_attribute
        col2 = attribute_clean.wafinfo_attribute
        request_size = len(col1) # request_size = 23
        wafinfo_size = len(col2) # wafinfo_size = 4
        for i in range(0, request_size):
                sheet.write(0, i, col1[i]) # 0~22
        sheet.write(0, request_size, 'location') # 23
        for i in range(0, wafinfo_size):
                sheet.write(0, request_size + i + 1, col2[i]) # 24~27
        wb.close()

# 清理不需要的列和空行
cwd = os.getcwd()
new_data_path = os.path.join(cwd, 'normal_data/normal_data_all_cleaned_12.xlsx') # 存储删除不必要后的数据
data_path = os.path.join(cwd, 'normal_data/normal_data_cleaned_12.xlsx') # 初步清洗后的表格
data = pd.read_excel(data_path)
data = data.drop([0]) # 删除标签下空白的行
data = data.drop(labels='rawquery', axis=1) # 删除缺失值过高的属性
data = data.drop(labels='agent', axis=1)
data = data.drop(labels='headers', axis=1)
data = data.drop(labels='Content-Length', axis=1)
data = data.drop(labels='Connection-Length', axis=1)
data = data.drop(labels='Content-Type', axis=1)
data = data.drop(labels='Keep-Alive', axis=1)
data = data.drop(labels='Accept', axis=1)
data = data.drop(labels='Connection', axis=1)
data.to_excel(new_data_path, sheet_name='report', index=False)

# 将表中空白的单元格填充为null
data = openpyxl.load_workbook(new_data_path)       # 打开目标Excel文件
sheet = data.active  # 打开该Excel里对应的sheet
for k in range(1,sheet.max_column+1):    # 对第1至X列单元格遍历
    for i in range(1,sheet.max_row+1):   # 对第1至X行单元格遍历
        if sheet.cell(row=i, column=k).value is None:    # 如果该单元格为空
            sheet.cell(i, k, 'null')     # 填入null
data.save(new_data_path)








