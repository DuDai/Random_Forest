'''
将12个normal数据集和1个attack数据集合并合并
'''
import openpyxl
import os

cwd = os.getcwd()
attack_data_final = os.path.join(cwd, 'attack_data/attack_data_final.xlsx')
normal_data_final_1 = os.path.join(cwd, 'normal_data/normal_data_final_1.xlsx')
normal_data_final_2 = os.path.join(cwd, 'normal_data/normal_data_final_2.xlsx')
normal_data_final_3 = os.path.join(cwd, 'normal_data/normal_data_final_3.xlsx')
normal_data_final_4 = os.path.join(cwd, 'normal_data/normal_data_final_4.xlsx')
normal_data_final_5 = os.path.join(cwd, 'normal_data/normal_data_final_5.xlsx')
normal_data_final_6 = os.path.join(cwd, 'normal_data/normal_data_final_6.xlsx')
normal_data_final_7 = os.path.join(cwd, 'normal_data/normal_data_final_7.xlsx')
normal_data_final_8 = os.path.join(cwd, 'normal_data/normal_data_final_8.xlsx')
normal_data_final_9 = os.path.join(cwd, 'normal_data/normal_data_final_9.xlsx')
normal_data_final_10 = os.path.join(cwd, 'normal_data/normal_data_final_10.xlsx')
normal_data_final_11 = os.path.join(cwd, 'normal_data/normal_data_final_11.xlsx')
normal_data_final_12 = os.path.join(cwd, 'normal_data/normal_data_final_12.xlsx')

# 文件列表
files_to_merge = [attack_data_final,normal_data_final_1,normal_data_final_2,normal_data_final_3,normal_data_final_4,normal_data_final_5,
             normal_data_final_6,normal_data_final_7,normal_data_final_8,normal_data_final_9,normal_data_final_10,normal_data_final_11,
             normal_data_final_12]

# 创建合并后的文件
wb_out = None
LINES_HEAD = 1
LINES_TOTAL = 1
count = 1

for files in files_to_merge:
    wb = openpyxl.load_workbook(files, read_only=True) # 打开待合并的文件
    sht = wb.active
    if wb_out is None: # 读入第一个表时全读
        wb_out = openpyxl.Workbook()  # 新建一个实例
        sht_out = wb_out.active  # 激活工作簿
        for rows in sht.iter_rows(min_row=LINES_HEAD, values_only=True):
            sht_out.append(rows)
    else:
        for rows in sht.iter_rows(min_row=LINES_HEAD+1, values_only=True):
            sht_out.append(rows)
    wb.close()
    print(count)
    count = count + 1

wb_out.save('./data_final.xlsx')
